import os
import shutil
import time
from copy import deepcopy
from pathlib import Path
from unittest import TestCase
from unittest.mock import Mock

import requests
from ingest.api.ingestapi import IngestApi
from ingest.importer.importer import XlsImporter
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.datastructures import FileStorage

from broker.service.spreadsheet_upload_service import SpreadsheetUploadService
from broker.submissions import ExportToSpreadsheetService

VALUE_ROW_NUMBER = 6
HEADER_ROW_NUMBER = 4


class BulkUpdateTest(TestCase):
    def setUp(self) -> None:
        # TODO add configuration to get ingest url from env variables
        # TODO currently it is hardcode for dev API

        jwt_token = os.environ.get('JWT_TOKEN')

        if not jwt_token:
            raise Exception('JWT token should be provided as an environment variable with the JWT_TOKEN variable.')

        self.ingest_url = "https://api.ingest.dev.archive.data.humancellatlas.org"
        self.ingest_api = IngestApi(self.ingest_url)
        self.ingest_api.set_token(jwt_token)

        self.export_service = ExportToSpreadsheetService(self.ingest_api)

        self.importer = XlsImporter(self.ingest_api)

        self.path_to_original_spreadsheet = '../resources/dcp_integration_test_metadata_1_SS2_bundle_orig.xlsx'
        self.path_to_spreadsheet_to_work_on = '../resources/dcp_integration_test_metadata_1_SS2_bundle.xlsx'

        storage_service = Mock()
        storage_service.store_submission_spreadsheet.return_value = self.path_to_spreadsheet_to_work_on
        self.spreadsheet_upload_svc = SpreadsheetUploadService(self.ingest_api, storage_service, self.importer)

        self.path_to_modified_spreadsheet = None
        self.modified_biomaterial_id = None

    def tearDown(self) -> None:
        self.__remove_temp_worksheets()

        self.__update_content('projects', self.project_id, self.original_project_content)
        self.__update_content('biomaterials', self.modified_biomaterial_id,
                              self.biomaterial_content_by_id.get(self.modified_biomaterial_id))
        self.__update_content('biomaterials', self.biomaterial_ids[0],
                              self.biomaterial_content_by_id.get(self.biomaterial_ids[0]))

    def test_export__modify__import_spreadsheet_flow(self):
        shutil.copyfile(self.path_to_original_spreadsheet, self.path_to_spreadsheet_to_work_on)

        # import a spreadsheet as a new submission
        sheet = Mock(FileStorage())
        sheet.filename = self.path_to_spreadsheet_to_work_on
        sheet.read = Mock()
        new_submission = \
            self.spreadsheet_upload_svc.async_upload(self.ingest_api.token, sheet, is_update=False)

        submission_uuid = new_submission['uuid']['uuid']
        submission = self.ingest_api.get_submission_by_uuid(submission_uuid)
        submission_id = self.__get_id_from_entity(submission)

        time.sleep(90)

        # get the original metadata
        original_project_payload = self.__get_entities_by_submission_id_and_type(submission_id, 'projects')[0]
        self.project_id = self.__get_id_from_entity(original_project_payload)
        self.original_project_content = original_project_payload.get('content')

        original_biomaterials_payload = self.__get_entities_by_submission_id_and_type(submission_id, 'biomaterials')
        self.biomaterial_ids = \
            [self.__get_id_from_entity(payload) for payload in original_biomaterials_payload]
        self.biomaterial_content_by_id = \
            {self.__get_id_from_entity(payload): payload.get('content')
             for payload in original_biomaterials_payload}

        # Modify project metadata on the UI (simulated by REST calls)
        project_content_with_ui_modification = deepcopy(self.original_project_content)
        updated_project_description = 'UI CHANGE ' + project_content_with_ui_modification.get('project_core').get(
            'project_description')
        project_content_with_ui_modification['project_core']['project_description'] = updated_project_description
        updated_contributor_name = 'UI CHANGE ' + project_content_with_ui_modification['contributors'][0]['name']
        project_content_with_ui_modification['contributors'][0]['name'] = updated_contributor_name
        self.__update_content('projects', self.project_id, project_content_with_ui_modification)

        biomaterial_content_with_ui_modification = deepcopy(list(self.biomaterial_content_by_id.values())[0])
        updated_bsd_accession = 'UI CHANGE ' + biomaterial_content_with_ui_modification['biomaterial_core']['biosamples_accession']
        biomaterial_content_with_ui_modification['biomaterial_core']['biosamples_accession'] = updated_bsd_accession
        self.__update_content('biomaterials', self.biomaterial_ids[0], biomaterial_content_with_ui_modification)

        # export a the modified submission (download a spreadsheet)
        # given a submission uuid /<submission_uuid>/spreadsheet
        self.path_to_modified_spreadsheet, spreadsheet = self.__export_spreadsheet(submission_uuid)

        # validate same cell values and then modify it then check it again that it has been modified
        project_sheet: Worksheet = spreadsheet.get_sheet_by_name('Project')
        updated_project_title = self.__update_project_title(project_sheet)

        specimen_sheet: Worksheet = spreadsheet.get_sheet_by_name('Specimen from organism')
        updated_biomaterial_name, self.modified_biomaterial_id = self.__update_biomaterial_name(specimen_sheet)

        self.__save_modified_spreadsheet(self.path_to_modified_spreadsheet, spreadsheet)

        # import the modified spreadsheet (submission)
        submission_url = self.ingest_url + '/submissionEnvelopes/' + submission_id
        self.importer.import_file(self.path_to_modified_spreadsheet, submission_url, is_update=True)

        # get the modified metadata
        updated_project_from_api = self.__get_entity_json_by_type_and_id('projects', self.project_id)
        updated_biomaterial1_from_api = \
            self.__get_entity_json_by_type_and_id('biomaterials', self.biomaterial_ids[0])
        updated_biomaterial2_from_api = \
            self.__get_entity_json_by_type_and_id('biomaterials', self.modified_biomaterial_id)

        # assert that the update has been applied on the metadata and check the older changes too
        self.assertEqual(
            updated_project_title, updated_project_from_api['content']['project_core']['project_title'])
        self.assertEqual(
            updated_biomaterial_name, updated_biomaterial2_from_api['content']['biomaterial_core']['biomaterial_name'])
        self.assertEqual(
            updated_project_description, updated_project_from_api['content']['project_core']['project_description'])
        self.assertEqual(
            updated_contributor_name, updated_project_from_api['content']['contributors'][0]['name'])
        self.assertEqual(
            updated_bsd_accession, updated_biomaterial1_from_api['content']['biomaterial_core']['biosamples_accession'])

    def __remove_temp_worksheets(self):
        downloaded_sheet_path = Path(self.path_to_modified_spreadsheet)
        downloaded_sheet_path.unlink()

        uploaded_sheet_path = Path(self.path_to_spreadsheet_to_work_on)
        uploaded_sheet_path.unlink()

    def __update_content(self, entity_type, entity_id, original_content):
        self.ingest_api.patch(self.ingest_url + f'/{entity_type}/' + entity_id, {'content': original_content})

    @staticmethod
    def __get_cell_by_header_name(worksheet: Worksheet, header_name):
        header_row = worksheet[HEADER_ROW_NUMBER]
        for cell in header_row:
            if header_name in cell.value:
                return cell.col_idx
        return None

    def __get_entities_by_submission_id_and_type(self, submission_id, entity_type):
        response = requests.get(self.ingest_url + f'/submissionEnvelopes/{submission_id}/{entity_type}').json()
        return response.get('_embedded').get(entity_type)

    def __get_entity_json_by_type_and_id(self, entity_type, entity_id):
        return requests.get(self.ingest_url + f'/{entity_type}/' + entity_id).json()

    def __export_spreadsheet(self, submission_uuid):
        spreadsheet = self.export_service.export(submission_uuid)
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        filename = f'{submission_uuid}_{timestamp}.xlsx'
        path_to_spreadsheet = 'temp_resources/' + filename
        spreadsheet.save(path_to_spreadsheet)

        return path_to_spreadsheet, spreadsheet

    def __update_project_title(self, project_sheet):
        project_title_column_index = self.__get_cell_by_header_name(project_sheet, 'project_title')
        orig_project_title = project_sheet.cell(VALUE_ROW_NUMBER, project_title_column_index).value
        updated_project_title = orig_project_title + ' SHEET UPDATE'
        project_sheet.cell(VALUE_ROW_NUMBER, project_title_column_index, updated_project_title)

        self.assertEqual(updated_project_title, project_sheet.cell(VALUE_ROW_NUMBER, project_title_column_index).value)

        return updated_project_title

    def __update_biomaterial_name(self, specimen_sheet):
        biomaterial_name_column_index = self.__get_cell_by_header_name(specimen_sheet, 'biomaterial_name')
        orig_biomaterial_name = specimen_sheet.cell(VALUE_ROW_NUMBER, biomaterial_name_column_index).value
        updated_biomaterial_name = orig_biomaterial_name + ' SHEET UPDATE'
        specimen_sheet.cell(VALUE_ROW_NUMBER, biomaterial_name_column_index, updated_biomaterial_name)

        self.assertEqual(updated_biomaterial_name, specimen_sheet.cell(VALUE_ROW_NUMBER, biomaterial_name_column_index).value)

        modified_biomaterial_id = self.__get_entity_id_by_sheet_and_entity_type(specimen_sheet, 'biomaterials')

        return updated_biomaterial_name, modified_biomaterial_id

    def __get_entity_id_by_sheet_and_entity_type(self, sheet, entity_type):
        uuid_column_index = self.__get_cell_by_header_name(sheet, 'uuid')
        entity_uuid = sheet.cell(VALUE_ROW_NUMBER, uuid_column_index).value
        modified_entity = self.ingest_api.get_entity_by_uuid(entity_type, entity_uuid)
        return self.__get_id_from_entity(modified_entity)

    @staticmethod
    def __get_id_from_entity(entity):
        return entity['_links']['self']['href'].split('/')[-1]

    @staticmethod
    def __save_modified_spreadsheet(path_to_spreadsheet, spreadsheet):
        spreadsheet.save(path_to_spreadsheet)
